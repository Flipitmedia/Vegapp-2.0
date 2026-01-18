"""
Sistema de Gestión de Pedidos - La Vega
Procesa exportaciones de Shopify y genera listas de compras y armado.
Diseñado e implementado por Flipit.media
"""

from fastapi import FastAPI, UploadFile, File, Request, Form, HTTPException
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import csv
import io
import re
from datetime import datetime, date, timedelta
from typing import Optional
import json
import os
from pathlib import Path

# Para generar Excel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Base de datos SQLite
import sqlite3

app = FastAPI(title="Sistema Gestión La Vega")

# Configurar archivos estáticos y templates
BASE_DIR = Path(__file__).resolve().parent
app.mount("/static", StaticFiles(directory=BASE_DIR / "static"), name="static")
templates = Jinja2Templates(directory=BASE_DIR / "templates")

# Directorio para archivos generados
OUTPUT_DIR = BASE_DIR / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)

DB_PATH = BASE_DIR / "vega.db"


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    """Inicializa la base de datos con las tablas necesarias."""
    conn = get_db()
    cursor = conn.cursor()
    
    # Tabla de categorías
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS categorias (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT UNIQUE NOT NULL,
            orden INTEGER DEFAULT 0
        )
    ''')
    
    # Tabla de mapeo producto -> categoría
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS producto_categoria (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            producto TEXT UNIQUE NOT NULL,
            categoria_id INTEGER,
            FOREIGN KEY (categoria_id) REFERENCES categorias(id)
        )
    ''')
    
    # Tabla de pedidos
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS pedidos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_number TEXT UNIQUE NOT NULL,
            email TEXT,
            comuna TEXT,
            fecha_entrega DATE,
            fecha_original DATE,
            direccion TEXT,
            telefono TEXT,
            nombre_cliente TEXT,
            total REAL,
            created_at TIMESTAMP,
            imported_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            status TEXT DEFAULT 'pendiente',
            completed_at TIMESTAMP
        )
    ''')
    
    # Agregar columna completed_at si no existe (para DBs existentes)
    try:
        cursor.execute("ALTER TABLE pedidos ADD COLUMN completed_at TIMESTAMP")
    except:
        pass  # Columna ya existe
    
    # Tabla de líneas de pedido
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS lineas_pedido (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pedido_id INTEGER,
            producto TEXT NOT NULL,
            cantidad INTEGER NOT NULL,
            precio REAL,
            sku TEXT,
            FOREIGN KEY (pedido_id) REFERENCES pedidos(id)
        )
    ''')
    
    # Tabla de configuración
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS configuracion (
            clave TEXT PRIMARY KEY,
            valor TEXT
        )
    ''')
    
    # Insertar categorías por defecto
    categorias_default = [
        ('Frutas', 1),
        ('Verduras', 2),
        ('Congelados', 3),
        ('Abarrotes', 4),
        ('Lácteos', 5),
        ('Carnes', 6),
        ('Otros', 99)
    ]
    
    for nombre, orden in categorias_default:
        cursor.execute('INSERT OR IGNORE INTO categorias (nombre, orden) VALUES (?, ?)', (nombre, orden))
    
    # Configuración por defecto
    config_default = [
        ('backup_email', ''),
        ('backup_frecuencia_dias', '3'),
        ('backup_hora', '08:00'),
        ('ultimo_backup', ''),
    ]
    
    for clave, valor in config_default:
        cursor.execute('INSERT OR IGNORE INTO configuracion (clave, valor) VALUES (?, ?)', (clave, valor))
    
    conn.commit()
    conn.close()


# Inicializar DB al arrancar
init_db()


def parse_note_attributes(note_attrs: str) -> dict:
    """Extrae comuna y fecha de entrega de los note attributes."""
    result = {'comuna': None, 'fecha_entrega': None}
    
    if not note_attrs:
        return result
    
    # Buscar comuna
    comuna_match = re.search(r'Comuna de Entrega:\s*([^\n]+)', note_attrs)
    if comuna_match:
        result['comuna'] = comuna_match.group(1).strip()
    
    # Buscar fecha
    fecha_match = re.search(r'Fecha de Entrega:\s*(\d{4}-\d{2}-\d{2})', note_attrs)
    if fecha_match:
        result['fecha_entrega'] = fecha_match.group(1)
    
    return result


def parse_shopify_csv(content: str) -> list:
    """Parsea el CSV de Shopify y agrupa por pedido."""
    reader = csv.DictReader(io.StringIO(content))
    
    orders = {}
    
    for row in reader:
        order_number = row.get('Name', '')
        if not order_number:
            continue
            
        if order_number not in orders:
            note_attrs = parse_note_attributes(row.get('Note Attributes', ''))
            
            created_at = None
            if row.get('Created at'):
                try:
                    created_at = datetime.strptime(
                        row['Created at'].split(' -')[0].split(' +')[0], 
                        '%Y-%m-%d %H:%M:%S'
                    )
                except:
                    pass
            
            orders[order_number] = {
                'order_number': order_number,
                'email': row.get('Email', ''),
                'comuna': note_attrs['comuna'],
                'fecha_entrega': note_attrs['fecha_entrega'],
                'nombre_cliente': row.get('Shipping Name', '') or row.get('Billing Name', ''),
                'direccion': row.get('Shipping Address1', ''),
                'telefono': row.get('Phone', '') or row.get('Shipping Phone', ''),
                'total': float(row.get('Total', 0) or 0),
                'created_at': created_at,
                'items': []
            }
        
        if row.get('Lineitem name'):
            orders[order_number]['items'].append({
                'producto': row['Lineitem name'],
                'cantidad': int(row.get('Lineitem quantity', 1) or 1),
                'precio': float(row.get('Lineitem price', 0) or 0),
                'sku': row.get('Lineitem sku', '')
            })
    
    return list(orders.values())


def get_config(clave: str) -> str:
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT valor FROM configuracion WHERE clave = ?", (clave,))
    row = cursor.fetchone()
    conn.close()
    return row[0] if row else ''


def set_config(clave: str, valor: str):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("INSERT OR REPLACE INTO configuracion (clave, valor) VALUES (?, ?)", (clave, valor))
    conn.commit()
    conn.close()


# ============================================
# RUTAS PRINCIPALES
# ============================================

@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    """Página principal."""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute("SELECT COUNT(*) FROM pedidos WHERE status = 'pendiente'")
    pedidos_pendientes = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM pedidos WHERE status = 'postergado'")
    pedidos_postergados = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(DISTINCT fecha_entrega) FROM pedidos WHERE status IN ('pendiente', 'postergado')")
    fechas_pendientes = cursor.fetchone()[0]
    
    hoy = date.today().isoformat()
    cursor.execute("SELECT COUNT(*) FROM pedidos WHERE fecha_entrega = ? AND status IN ('pendiente', 'postergado')", (hoy,))
    pedidos_hoy = cursor.fetchone()[0]
    
    cursor.execute('''
        SELECT COUNT(DISTINCT lp.producto) 
        FROM lineas_pedido lp
        LEFT JOIN producto_categoria pc ON lp.producto = pc.producto
        WHERE pc.id IS NULL
    ''')
    sin_categoria = cursor.fetchone()[0]
    
    conn.close()
    
    return templates.TemplateResponse("index.html", {
        "request": request,
        "pedidos_pendientes": pedidos_pendientes,
        "pedidos_postergados": pedidos_postergados,
        "fechas_pendientes": fechas_pendientes,
        "pedidos_hoy": pedidos_hoy,
        "sin_categoria": sin_categoria,
        "fecha_hoy": hoy
    })


@app.post("/upload")
async def upload_csv(file: UploadFile = File(...)):
    """Sube y procesa un CSV de Shopify."""
    if not file.filename.endswith('.csv'):
        raise HTTPException(400, "El archivo debe ser CSV")
    
    content = await file.read()
    content = content.decode('utf-8-sig')
    
    orders = parse_shopify_csv(content)
    
    conn = get_db()
    cursor = conn.cursor()
    
    nuevos = 0
    duplicados = 0
    sin_fecha = 0
    
    for order in orders:
        cursor.execute("SELECT id FROM pedidos WHERE order_number = ?", (order['order_number'],))
        existing = cursor.fetchone()
        
        if existing:
            duplicados += 1
            continue
        
        if not order['fecha_entrega']:
            sin_fecha += 1
            continue
        
        cursor.execute('''
            INSERT INTO pedidos (order_number, email, comuna, fecha_entrega, fecha_original, direccion, telefono, nombre_cliente, total, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            order['order_number'],
            order['email'],
            order['comuna'],
            order['fecha_entrega'],
            order['fecha_entrega'],
            order['direccion'],
            order['telefono'],
            order['nombre_cliente'],
            order['total'],
            order['created_at']
        ))
        
        pedido_id = cursor.lastrowid
        
        for item in order['items']:
            cursor.execute('''
                INSERT INTO lineas_pedido (pedido_id, producto, cantidad, precio, sku)
                VALUES (?, ?, ?, ?, ?)
            ''', (pedido_id, item['producto'], item['cantidad'], item['precio'], item['sku']))
        
        nuevos += 1
    
    conn.commit()
    conn.close()
    
    return {
        "success": True,
        "nuevos": nuevos,
        "duplicados": duplicados,
        "sin_fecha": sin_fecha,
        "total": len(orders)
    }


@app.get("/api/categorias")
async def get_categorias():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, nombre, orden FROM categorias ORDER BY orden")
    categorias = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return categorias


@app.post("/api/categorias")
async def create_categoria(nombre: str = Form(...)):
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT MAX(orden) FROM categorias")
        max_orden = cursor.fetchone()[0] or 0
        cursor.execute("INSERT INTO categorias (nombre, orden) VALUES (?, ?)", (nombre, max_orden + 1))
        conn.commit()
        return {"success": True, "id": cursor.lastrowid}
    except sqlite3.IntegrityError:
        raise HTTPException(400, "La categoría ya existe")
    finally:
        conn.close()


@app.get("/api/productos-sin-categoria")
async def get_productos_sin_categoria():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT DISTINCT lp.producto
        FROM lineas_pedido lp
        LEFT JOIN producto_categoria pc ON lp.producto = pc.producto
        WHERE pc.id IS NULL
        ORDER BY lp.producto
    ''')
    productos = [row[0] for row in cursor.fetchall()]
    conn.close()
    return productos


@app.post("/api/asignar-categoria")
async def asignar_categoria(producto: str = Form(...), categoria_id: int = Form(...)):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('INSERT OR REPLACE INTO producto_categoria (producto, categoria_id) VALUES (?, ?)', (producto, categoria_id))
    conn.commit()
    conn.close()
    return {"success": True}


@app.get("/api/pedidos")
async def get_pedidos(fecha: Optional[str] = None, status: Optional[str] = None):
    conn = get_db()
    cursor = conn.cursor()
    
    query = "SELECT * FROM pedidos WHERE 1=1"
    params = []
    
    if fecha:
        query += " AND fecha_entrega = ?"
        params.append(fecha)
    
    if status:
        if status == 'activo':
            query += " AND status IN ('pendiente', 'postergado')"
        else:
            query += " AND status = ?"
            params.append(status)
    
    query += " ORDER BY fecha_entrega, order_number"
    
    cursor.execute(query, params)
    pedidos = [dict(row) for row in cursor.fetchall()]
    
    for pedido in pedidos:
        cursor.execute("SELECT * FROM lineas_pedido WHERE pedido_id = ?", (pedido['id'],))
        pedido['items'] = [dict(row) for row in cursor.fetchall()]
    
    conn.close()
    return pedidos


@app.get("/api/fechas-pendientes")
async def get_fechas_pendientes():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT fecha_entrega, COUNT(*) as cantidad,
               SUM(CASE WHEN status = 'postergado' THEN 1 ELSE 0 END) as postergados
        FROM pedidos 
        WHERE status IN ('pendiente', 'postergado')
        GROUP BY fecha_entrega
        ORDER BY fecha_entrega
    ''')
    fechas = [{"fecha": row[0], "cantidad": row[1], "postergados": row[2]} for row in cursor.fetchall()]
    conn.close()
    return fechas


@app.get("/api/lista-compras/{fecha}")
async def get_lista_compras(fecha: str):
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT 
            lp.producto,
            SUM(lp.cantidad) as cantidad_total,
            COALESCE(c.nombre, 'Sin Categoría') as categoria,
            COALESCE(c.orden, 999) as categoria_orden
        FROM lineas_pedido lp
        JOIN pedidos p ON lp.pedido_id = p.id
        LEFT JOIN producto_categoria pc ON lp.producto = pc.producto
        LEFT JOIN categorias c ON pc.categoria_id = c.id
        WHERE p.fecha_entrega = ? AND p.status IN ('pendiente', 'postergado')
        GROUP BY lp.producto
        ORDER BY categoria_orden, c.nombre, lp.producto
    ''', (fecha,))
    
    items = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    por_categoria = {}
    for item in items:
        cat = item['categoria']
        if cat not in por_categoria:
            por_categoria[cat] = []
        por_categoria[cat].append({
            'producto': item['producto'],
            'cantidad': item['cantidad_total']
        })
    
    return por_categoria


@app.get("/descargar/lista-compras/{fecha}")
async def descargar_lista_compras(fecha: str):
    lista = await get_lista_compras(fecha)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Lista de Compras"
    
    header_fill = PatternFill(start_color="2E5C46", end_color="2E5C46", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    cat_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    cat_font = Font(bold=True, size=11, color="2E5C46")
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Título
    ws.merge_cells('A1:C1')
    ws['A1'] = f"🥬 Lista de Compras - {fecha}"
    ws['A1'].font = Font(bold=True, size=16, color="2E5C46")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    ws['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(italic=True, color="666666")
    
    # Headers
    ws['A4'] = "Producto"
    ws['B4'] = "Cantidad"
    ws['C4'] = "✓"
    for col in ['A', 'B', 'C']:
        ws[f'{col}4'].font = header_font
        ws[f'{col}4'].fill = header_fill
        ws[f'{col}4'].border = border
        ws[f'{col}4'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[4].height = 25
    
    row = 5
    for categoria, productos in lista.items():
        # Categoría
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = f"📦 {categoria}"
        ws[f'A{row}'].font = cat_font
        ws[f'A{row}'].fill = cat_fill
        ws[f'A{row}'].border = border
        ws.row_dimensions[row].height = 22
        row += 1
        
        for prod in productos:
            ws[f'A{row}'] = prod['producto']
            ws[f'B{row}'] = prod['cantidad']
            ws[f'C{row}'] = "☐"
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = border
            ws[f'B{row}'].alignment = Alignment(horizontal='center')
            ws[f'C{row}'].alignment = Alignment(horizontal='center')
            ws[f'C{row}'].font = Font(size=14)
            row += 1
        
        row += 1  # Espacio entre categorías
    
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 8
    
    filename = f"lista_compras_{fecha}.xlsx"
    filepath = OUTPUT_DIR / filename
    wb.save(filepath)
    
    return FileResponse(filepath, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/descargar/pedidos-armado/{fecha}")
async def descargar_pedidos_armado(fecha: str):
    pedidos = await get_pedidos(fecha=fecha, status='activo')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos Armado"
    
    header_fill = PatternFill(start_color="2E5C46", end_color="2E5C46", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    order_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    order_font = Font(bold=True, size=12, color="2E5C46")
    postergado_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Título
    ws.merge_cells('A1:D1')
    ws['A1'] = f"📦 Pedidos para Armar - {fecha}"
    ws['A1'].font = Font(bold=True, size=16, color="2E5C46")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    ws['A2'] = f"Total: {len(pedidos)} pedidos | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(italic=True, color="666666")
    
    row = 4
    for pedido in pedidos:
        # Header del pedido
        ws.merge_cells(f'A{row}:D{row}')
        status_emoji = "⏳" if pedido['status'] == 'postergado' else "📋"
        ws[f'A{row}'] = f"{status_emoji} {pedido['order_number']} | {pedido['nombre_cliente'] or 'Sin nombre'} | {pedido['comuna'] or 'Sin comuna'}"
        ws[f'A{row}'].font = order_font
        ws[f'A{row}'].fill = postergado_fill if pedido['status'] == 'postergado' else order_fill
        ws[f'A{row}'].border = border
        ws.row_dimensions[row].height = 28
        row += 1
        
        if pedido['direccion']:
            ws[f'A{row}'] = f"📍 {pedido['direccion']}"
            ws[f'A{row}'].font = Font(italic=True, color="666666", size=10)
            row += 1
        
        # Headers de productos
        ws[f'A{row}'] = "Producto"
        ws[f'B{row}'] = "Cant."
        ws[f'C{row}'] = "✓"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].border = border
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
        row += 1
        
        for item in pedido['items']:
            ws[f'A{row}'] = item['producto']
            ws[f'B{row}'] = item['cantidad']
            ws[f'C{row}'] = "☐"
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = border
            ws[f'B{row}'].alignment = Alignment(horizontal='center')
            ws[f'C{row}'].alignment = Alignment(horizontal='center')
            ws[f'C{row}'].font = Font(size=14)
            row += 1
        
        row += 2  # Espacio entre pedidos
    
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 8
    
    filename = f"pedidos_armado_{fecha}.xlsx"
    filepath = OUTPUT_DIR / filename
    wb.save(filepath)
    
    return FileResponse(filepath, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.post("/api/pedidos/{pedido_id}/completar")
async def completar_pedido(pedido_id: int):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE pedidos 
        SET status = 'completado', 
            completed_at = CURRENT_TIMESTAMP 
        WHERE id = ?
    """, (pedido_id,))
    conn.commit()
    conn.close()
    return {"success": True}


@app.post("/api/pedidos/{pedido_id}/reactivar")
async def reactivar_pedido(pedido_id: int):
    """Deshace el completado de un pedido, volviéndolo a pendiente."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE pedidos 
        SET status = 'pendiente', 
            completed_at = NULL 
        WHERE id = ?
    """, (pedido_id,))
    conn.commit()
    conn.close()
    return {"success": True}


@app.get("/api/pedidos-completados")
async def get_pedidos_completados(limit: int = 50):
    """Obtiene los últimos pedidos completados."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, order_number, nombre_cliente, comuna, fecha_entrega, completed_at, total
        FROM pedidos 
        WHERE status = 'completado'
        ORDER BY completed_at DESC
        LIMIT ?
    """, (limit,))
    pedidos = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return pedidos


@app.post("/api/auto-completar-pasados")
async def auto_completar_pasados():
    """Auto-completa pedidos con fecha de entrega pasada."""
    conn = get_db()
    cursor = conn.cursor()
    hoy = date.today().isoformat()
    
    # Contar cuántos se van a completar
    cursor.execute("""
        SELECT COUNT(*) FROM pedidos 
        WHERE fecha_entrega < ? AND status = 'pendiente'
    """, (hoy,))
    cantidad = cursor.fetchone()[0]
    
    if cantidad > 0:
        cursor.execute("""
            UPDATE pedidos 
            SET status = 'completado', 
                completed_at = CURRENT_TIMESTAMP 
            WHERE fecha_entrega < ? AND status = 'pendiente'
        """, (hoy,))
        conn.commit()
    
    conn.close()
    return {"success": True, "completados": cantidad}


@app.get("/api/pedidos-pasados-pendientes")
async def get_pedidos_pasados_pendientes():
    """Obtiene pedidos con fecha pasada que aún están pendientes."""
    conn = get_db()
    cursor = conn.cursor()
    hoy = date.today().isoformat()
    cursor.execute("""
        SELECT COUNT(*) FROM pedidos 
        WHERE fecha_entrega < ? AND status = 'pendiente'
    """, (hoy,))
    cantidad = cursor.fetchone()[0]
    conn.close()
    return {"cantidad": cantidad}


@app.post("/api/pedidos/{pedido_id}/postergar")
async def postergar_pedido(pedido_id: int, nueva_fecha: str = Form(...)):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE pedidos SET fecha_entrega = ?, status = 'postergado' WHERE id = ?", (nueva_fecha, pedido_id))
    conn.commit()
    conn.close()
    return {"success": True}


@app.delete("/api/pedidos/{pedido_id}")
async def eliminar_pedido(pedido_id: int):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM lineas_pedido WHERE pedido_id = ?", (pedido_id,))
    cursor.execute("DELETE FROM pedidos WHERE id = ?", (pedido_id,))
    conn.commit()
    conn.close()
    return {"success": True}


# ============================================
# BACKUP
# ============================================

def generar_backup_excel() -> Path:
    conn = get_db()
    cursor = conn.cursor()
    
    wb = Workbook()
    
    # Hoja Pedidos
    ws = wb.active
    ws.title = "Pedidos"
    headers = ['ID', 'Número', 'Email', 'Comuna', 'Fecha Entrega', 'Fecha Original', 'Dirección', 'Teléfono', 'Cliente', 'Total', 'Creado', 'Importado', 'Estado']
    ws.append(headers)
    header_fill = PatternFill(start_color="2E5C46", end_color="2E5C46", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).fill = header_fill
        ws.cell(row=1, column=col).font = header_font
    cursor.execute("SELECT * FROM pedidos")
    for row in cursor.fetchall():
        ws.append(list(row))
    
    # Hoja Líneas
    ws2 = wb.create_sheet("Lineas")
    ws2.append(['ID', 'Pedido ID', 'Producto', 'Cantidad', 'Precio', 'SKU'])
    for col in range(1, 7):
        ws2.cell(row=1, column=col).fill = header_fill
        ws2.cell(row=1, column=col).font = header_font
    cursor.execute("SELECT * FROM lineas_pedido")
    for row in cursor.fetchall():
        ws2.append(list(row))
    
    # Hoja Categorías
    ws3 = wb.create_sheet("Categorias")
    ws3.append(['ID', 'Nombre', 'Orden'])
    for col in range(1, 4):
        ws3.cell(row=1, column=col).fill = header_fill
        ws3.cell(row=1, column=col).font = header_font
    cursor.execute("SELECT * FROM categorias")
    for row in cursor.fetchall():
        ws3.append(list(row))
    
    # Hoja Producto-Categoría
    ws4 = wb.create_sheet("ProductoCategoria")
    ws4.append(['ID', 'Producto', 'Categoria ID'])
    for col in range(1, 4):
        ws4.cell(row=1, column=col).fill = header_fill
        ws4.cell(row=1, column=col).font = header_font
    cursor.execute("SELECT * FROM producto_categoria")
    for row in cursor.fetchall():
        ws4.append(list(row))
    
    conn.close()
    
    fecha_str = datetime.now().strftime("%Y-%m-%d_%H%M")
    filename = f"backup_vega_{fecha_str}.xlsx"
    filepath = OUTPUT_DIR / filename
    wb.save(filepath)
    
    return filepath


@app.get("/descargar/backup")
async def descargar_backup():
    filepath = generar_backup_excel()
    return FileResponse(filepath, filename=filepath.name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.post("/api/backup/restaurar")
async def restaurar_backup(file: UploadFile = File(...), auto_completar: bool = Form(True)):
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(400, "Debe ser archivo Excel (.xlsx)")
    
    temp_path = OUTPUT_DIR / f"temp_{datetime.now().timestamp()}.xlsx"
    content = await file.read()
    with open(temp_path, 'wb') as f:
        f.write(content)
    
    try:
        wb = load_workbook(temp_path)
        conn = get_db()
        cursor = conn.cursor()
        
        stats = {'pedidos': 0, 'lineas': 0, 'auto_completados': 0}
        hoy = date.today().isoformat()
        
        # Restaurar pedidos
        if "Pedidos" in wb.sheetnames:
            ws = wb["Pedidos"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    status = row[12] if row[12] else 'pendiente'
                    fecha = str(row[4]) if row[4] else None
                    
                    if auto_completar and fecha and fecha < hoy and status == 'pendiente':
                        status = 'completado'
                        stats['auto_completados'] += 1
                    
                    cursor.execute('''
                        INSERT OR REPLACE INTO pedidos 
                        (id, order_number, email, comuna, fecha_entrega, fecha_original, direccion, telefono, nombre_cliente, total, created_at, imported_at, status)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (row[0], row[1], row[2], row[3], fecha, row[5], row[6], row[7], row[8], row[9], row[10], row[11], status))
                    stats['pedidos'] += 1
        
        # Restaurar líneas
        if "Lineas" in wb.sheetnames:
            ws = wb["Lineas"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    cursor.execute('INSERT OR REPLACE INTO lineas_pedido (id, pedido_id, producto, cantidad, precio, sku) VALUES (?, ?, ?, ?, ?, ?)', row)
                    stats['lineas'] += 1
        
        conn.commit()
        conn.close()
        
        return {"success": True, "estadisticas": stats}
    finally:
        if temp_path.exists():
            temp_path.unlink()


@app.get("/api/backup/config")
async def get_backup_config():
    return {
        "email": get_config('backup_email'),
        "frecuencia_dias": int(get_config('backup_frecuencia_dias') or 3),
        "hora": get_config('backup_hora'),
        "ultimo_backup": get_config('ultimo_backup')
    }


@app.post("/api/backup/config")
async def set_backup_config(email: str = Form(''), frecuencia_dias: int = Form(3), hora: str = Form('08:00')):
    set_config('backup_email', email)
    set_config('backup_frecuencia_dias', str(frecuencia_dias))
    set_config('backup_hora', hora)
    return {"success": True}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
